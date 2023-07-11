from urllib.request import Request, urlopen as uReq
from bs4 import BeautifulSoup as soup
from selenium import webdriver
import time
from selenium.webdriver.chrome.options import Options
from openpyxl.styles import PatternFill
from webdriver_manager.chrome import ChromeDriverManager
import requests
import io
from django.conf import settings
import time
import json
from WebApp.models.Player import Player
from WebApp.models.Position import Position
from WebApp.models.Team import Team
from WebApp.models.Nation import Nation
from WebApp.models.League import League


def get_id_from_href(link):
    start = link.find("player/")
    end = link.find('/', start + 8)
    return link[start + 7:end]

def get_player_index(obj, position):
    players = list(Player.objects.filter(position=position))
    for player in players:
        if obj.name == player.name:
            return player.id
    return -1


def strip_player_positions_string():
    players = Player.objects.all()
    for player in players:
        player.specific_position = player.specific_position.split('|')[0].strip()
        player.save()


def get_players_db():
    positions = {'Goalkeepers': 'GK', 'Center Backs': 'CB', 'Full Backs': 'RB,LB', 'Defensive Midfielders': 'CDM,CM',
                 'Ofensive Midfielders': 'CAM', 'Wingers': 'LW,LF,LM,RF,RW,RM', 'Attackers': 'ST,CF'}

    base_url = "https://www.fifacm.com/players?position="

    option = Options()
    # option.headless = True
    options = webdriver.ChromeOptions()
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    options.add_argument("--disable-blink-features=AutomationControlled")
    driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)
    # driver.maximize_window()

    for name, positions in positions.items():
        query = Position.objects.filter(name=name)
        if len(query) == 0:
            position = Position.create(name, positions)
        else:
            position = query[0]

        for pagina in range(1, 4):
            url = base_url + positions

            url = url + '&page=' + str(pagina)
            print(url)

            driver.get(url)

            driver.maximize_window()
            time.sleep(5)

            igs_btns = driver.find_elements_by_class_name("igs-btn")
            scroll = 0
            cont = 0
            scrollStride = 450
            if positions == 'GK':
                scrollStride = 210

            for btn in igs_btns:
                btn.click()
                driver.execute_script("window.scrollTo({}, {})".format(
                    scroll, scroll + scrollStride))
                scroll += scrollStride
                time.sleep(1)

                element = driver.find_elements_by_class_name(
                    "site-players-page")[0]
                htmlContent = element.get_attribute('outerHTML')

                page_soup = soup(htmlContent, "html.parser")

                table = page_soup.find("table")
                trList = table.findAll("tr")[1+cont:1+cont+2]

                # Basic Info
                player = Player(position=position)

                tds = trList[0].findAll("td")
                player.id = tds[0].find(
                    'div', {"class": "igs-btn"})['data-playerid']

                other_player = Player.objects.filter(id=player.id)
                if other_player.exists():
                    cont += 2
                    continue

                team = tds[1].findAll(
                    'img', {"class": "team-img"})[0]['data-original-title']
                nation = tds[1].findAll(
                    'img', {"class": "team-img"})[1]['data-original-title']
                team_img = tds[1].findAll(
                    'img', {"class": "team-img"})[0]['src']
                nation_img = tds[1].findAll(
                    'img', {"class": "team-img"})[1]['src']

                query = Team.objects.filter(name=team)
                if len(query) == 0:
                    if('notfound' not in team_img):
                        image_content = requests.get(team_img).content
                        image_file = io.BytesIO(image_content)
                        file_path = settings.MEDIA_ROOT + \
                            "/Teams/{}.jpg".format(team.replace(' ', ''))
                        open(file_path, "wb").write(image_file.getbuffer())

                        team_img = file_path
                    team = Team.objects.create(name=team, image_path=team_img)
                else:
                    team = query[0]
                player.team_origin = team

                query = Nation.objects.filter(name=nation)
                if len(query) == 0:
                    if('notfound' not in nation_img):
                        image_content = requests.get(nation_img).content
                        image_file = io.BytesIO(image_content)
                        file_path = settings.MEDIA_ROOT + \
                            "/Nations/{}.jpg".format(nation.replace(' ', ''))
                        open(file_path, "wb").write(image_file.getbuffer())

                        nation_img = file_path
                    nation = Nation.create(nation, nation_img)
                else:
                    nation = query[0]
                player.nation = nation

                player.specific_position = ''.join(tds[1].find(
                    'div', {"class": "player-position-cln"}).findAll(text=True)).split(',')[0].split('|')[0].strip()
                stats = trList[1].find(
                    "div", {"class": "player-stats"}).findAll("div", {"class": "col-md-2"})
                player.name = ''.join(tds[1].find(
                    'div', {"class": "player-name"}).find("a").findAll(text=True))

                player_img_url = tds[1].find(
                    'img', {"class": "player-img-info"})['src']
                if('notfound_player' not in player_img_url):
                    image_content = requests.get(player_img_url).content
                    image_file = io.BytesIO(image_content)
                    file_path = settings.MEDIA_ROOT + \
                        "/Players/{}_{}.jpg".format(
                            player.name.replace(' ', ''), player.id)
                    open(file_path, "wb").write(image_file.getbuffer())

                    player.image_path = file_path
                else:
                    cont += 2
                    continue
                player.overall = ''.join(tds[2].find(
                    'div', {"class": "player-overall"}).findAll(text=True))
                player.pace = ''.join(stats[0].find(
                    "div", {"class": "main-stat-rating-title"}).findAll(text=True))
                player.shooting = ''.join(stats[1].find(
                    "div", {"class": "main-stat-rating-title"}).findAll(text=True))
                player.passing = ''.join(stats[2].find(
                    "div", {"class": "main-stat-rating-title"}).findAll(text=True))
                player.dribbling = ''.join(stats[3].find(
                    "div", {"class": "main-stat-rating-title"}).findAll(text=True))
                player.defending = ''.join(stats[4].find(
                    "div", {"class": "main-stat-rating-title"}).findAll(text=True))
                player.physical = ''.join(stats[5].find(
                    "div", {"class": "main-stat-rating-title"}).findAll(text=True))

                player.save()

                cont += 2

    driver.quit()

    update_teams_leagues()

def update_teams_leagues():
    f = open('leagues.json', encoding="utf-8")
    leagues = json.load(f)
    for league, teams in leagues.items():
        league_obj = League.objects.filter(name=league)
        if len(league_obj) > 0:
            league_obj = league_obj[0]
        else:
            league_obj = League.create(name=league)
        for team in teams:
            try:
                team_obj = Team.objects.get(name=team)
                team_obj.league = league_obj
                team_obj.save()
            except:
                continue

def get_players_by_position_algorithm(position_id, n):
    n = int(n)
    position_obj = Position.objects.get(id=position_id)
    players = list(Player.objects.all().order_by('-overall'))
    players = [player for player in players if player.specific_position in position_obj.specific_positions.split(',')]
    players.sort(key=lambda item: (item.overall, item.pace), reverse=True)
    players = players[:n]
    return players


def colour_cell(cell, stat):
    red_fill = PatternFill(start_color='FF0000',
                          end_color='FF0000', fill_type='solid')
    orange_fill = PatternFill(start_color='FF8F00',
                             end_color='FF8F00', fill_type='solid')
    yellow_fill = PatternFill(start_color='F6EB0A',
                             end_color='F6EB0A', fill_type='solid')
    light_green_fill = PatternFill(
        start_color='2DD63A', end_color='2DD63A', fill_type='solid')
    green_fill = PatternFill(start_color='3E9C45',
                            end_color='3E9C45', fill_type='solid')
    cell.value = stat

    if stat >= 90:
        cell.fill = green_fill
    elif stat >= 80:
        cell.fill = light_green_fill
    elif stat >= 70:
        cell.fill = yellow_fill
    elif stat >= 60:
        cell.fill = orange_fill
    else:
        cell.fill = red_fill

    return cell


def get_players_by_position_string(sheet, position, n):
    players = get_players_by_position_algorithm(position.id, n)
    write_players_sheet(sheet, players)
    return sheet


def write_players_sheet(sheet, players):
    sheet.oddHeader.center.size = 14
    sheet.oddHeader.center.font = "Calibri,Bold"
    sheet.cell(row=1, column=1).value = 'PLAYER'
    sheet.cell(row=1, column=2).value = 'TEAM'
    sheet.cell(row=1, column=3).value = 'NATION'
    sheet.cell(row=1, column=4).value = 'POSITION'
    sheet.cell(row=1, column=5).value = 'OVERALL'
    sheet.cell(row=1, column=6).value = 'PACE'
    sheet.cell(row=1, column=7).value = 'SHOOTING'
    sheet.cell(row=1, column=8).value = 'PASSING'
    sheet.cell(row=1, column=9).value = 'DRIBBLING'
    sheet.cell(row=1, column=10).value = 'DEFENDING'
    sheet.cell(row=1, column=11).value = 'PHYSICAL'
    sheet.cell(row=1, column=12).value = 'VALUE'

    for i, player in enumerate(players):
        sheet.cell(row=i+2, column=1).value = player.name
        sheet.cell(row=i+2, column=2).value = player.team_origin.name
        sheet.cell(row=i+2, column=3).value = player.nation.name
        sheet.cell(row=i+2, column=4).value = player.specific_position
        sheet.cell(row=i+2, column=5).value = player.overall
        colour_cell(sheet.cell(row=i+2, column=6), player.pace)
        colour_cell(sheet.cell(row=i+2, column=7), player.shooting)
        colour_cell(sheet.cell(row=i+2, column=8), player.passing)
        colour_cell(sheet.cell(row=i+2, column=9), player.dribbling)
        colour_cell(sheet.cell(row=i+2, column=10), player.defending)
        colour_cell(sheet.cell(row=i+2, column=11), player.physical)

        # if player.value != None:
        sheet.cell(row=i+2, column=12).value = player.value

    return sheet
