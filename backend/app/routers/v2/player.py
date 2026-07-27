import os
import typing as t
from io import BytesIO

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from pydantic import BaseModel
from sqlalchemy import select, func, and_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.auth.users import current_active_user
from app.db.base import get_async_session
from app.models_orm.championship import Championship
from app.models_orm.draft_pick import DraftPick
from app.models_orm.participant import Participant
from app.models_orm.player import Player
from app.models_orm.position import Position
from app.models_orm.team import Team
from app.models_orm.nation import Nation
from app.models_orm.user import User
from app.utils.export import write_players_sheet, write_transfers_sheet

_limit_by_position = {
    "Goalkeepers": 1,
    "Center Backs": 3,
    "Full Backs": 2.5,
    "Defensive Midfielders": 3,
    "Ofensive Midfielders": 1.5,
    "Wingers": 2.5,
    "Attackers": 2,
}

router = APIRouter()


class BuyPlayerModel(BaseModel):
    participant_id: t.Optional[int] = None
    value: t.Optional[int] = None


class PlayerResponse(BaseModel):
    id: str
    position_id: int
    position_name: str
    team_participant: t.Optional[str] = None
    team_participant_name: t.Optional[str] = None
    team_participant_image_path: t.Optional[str] = None
    participant_id: t.Optional[int] = None
    team_origin: str
    team_origin_name: str
    team_origin_image_path: str
    name: str
    nation: str
    nation_name: str
    nation_image_path: str
    specific_position: str
    overall: int
    pace: int
    shooting: int
    passing: int
    dribbling: int
    defending: int
    physical: int
    image_path: str
    value: t.Optional[int] = None

    class Config:
        from_attributes = True


def _player_to_response(player: Player, draft_pick: t.Optional[DraftPick] = None, participant: t.Optional[Participant] = None) -> PlayerResponse:
    team_participant = None
    team_participant_name = None
    team_participant_image_path = None
    participant_id = None
    value = None

    if draft_pick and participant:
        team_participant = participant.team_name
        if participant.team:
            team_participant_name = participant.team.name
            team_participant_image_path = participant.team.image_path
        participant_id = participant.id
        value = draft_pick.value

    return PlayerResponse(
        id=player.id,
        position_id=player.position_id,
        position_name=player.position.name if player.position else "",
        team_participant=team_participant,
        team_participant_name=team_participant_name,
        team_participant_image_path=team_participant_image_path,
        participant_id=participant_id,
        team_origin=player.team_origin,
        team_origin_name=player.team_origin_rel.name if player.team_origin_rel else player.team_origin,
        team_origin_image_path=player.team_origin_rel.image_path if player.team_origin_rel else "",
        name=player.name,
        nation=player.nation,
        nation_name=player.nation_rel.name if player.nation_rel else player.nation,
        nation_image_path=player.nation_rel.image_path if player.nation_rel else "",
        specific_position=player.specific_position,
        overall=player.overall,
        pace=player.pace,
        shooting=player.shooting,
        passing=player.passing,
        dribbling=player.dribbling,
        defending=player.defending,
        physical=player.physical,
        image_path=player.image_path,
        value=value,
    )


def _get_draft_player_ids(position_name: str = None) -> t.List[str]:
    """Read player IDs from sheet.xlsx for the draft pool."""
    try:
        possible_paths = [
            "/app/sheet.xlsx",
            os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "sheet.xlsx"),
            os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))), "sheet.xlsx"),
        ]

        sheet_path = None
        for path in possible_paths:
            if os.path.exists(path):
                sheet_path = path
                break

        if not sheet_path:
            return []

        workbook = openpyxl.load_workbook(sheet_path)
        player_names = []

        if position_name:
            worksheet = None
            for sheet_name in workbook.sheetnames:
                if sheet_name.lower() == position_name.lower():
                    worksheet = workbook[sheet_name]
                    break
            if not worksheet:
                workbook.close()
                return []
            for row in worksheet.iter_rows(min_row=2, max_col=1, values_only=True):
                if row[0] and str(row[0]).strip():
                    player_names.append(str(row[0]).strip())
        else:
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                for row in worksheet.iter_rows(min_row=2, max_col=1, values_only=True):
                    if row[0] and str(row[0]).strip():
                        player_names.append(str(row[0]).strip())

        workbook.close()
        return player_names
    except Exception as e:
        print(f"Error reading sheet.xlsx: {e}")
        return []


@router.get("/{championship_id}/player/list", response_model=t.List[PlayerResponse])
async def list_players(
    championship_id: int,
    position: t.Optional[int] = Query(None, description="Filter by position ID"),
    participant_id: t.Optional[int] = Query(None, description="Filter by participant (drafted players)"),
    draft_only: bool = Query(True, description="Only include players from sheet.xlsx"),
    user: User = Depends(current_active_user),
    session: AsyncSession = Depends(get_async_session),
):
    # Verify championship ownership
    champ_stmt = select(Championship).where(
        Championship.id == championship_id, Championship.owner_id == str(user.id)
    )
    champ = (await session.execute(champ_stmt)).scalar_one_or_none()
    if not champ:
        raise HTTPException(status_code=404, detail="Championship not found.")

    # Count participants for limit calculation
    p_count_stmt = select(func.count(Participant.id)).where(Participant.championship_id == championship_id)
    num_participants = (await session.execute(p_count_stmt)).scalar() or 0

    # Build player query
    stmt = select(Player).options(
        selectinload(Player.position),
        selectinload(Player.team_origin_rel),
        selectinload(Player.nation_rel),
    )

    position_name = None
    if position:
        pos_stmt = select(Position).where(Position.id == position)
        pos_obj = (await session.execute(pos_stmt)).scalar_one_or_none()
        if not pos_obj:
            raise HTTPException(status_code=404, detail="Position not found.")
        position_name = pos_obj.name
        stmt = stmt.where(Player.position_id == position)

    # Filter by draft sheet
    if draft_only:
        draft_names = _get_draft_player_ids(position_name)
        if draft_names:
            stmt = stmt.where(Player.name.in_(draft_names))
        else:
            return []

    stmt = stmt.order_by(Player.overall.desc(), Player.pace.desc())

    # Apply limit
    if position and num_participants and position_name in _limit_by_position:
        limit = int(num_participants * _limit_by_position[position_name])
        stmt = stmt.limit(limit)

    result = await session.execute(stmt)
    players = result.scalars().all()

    # Get all draft picks for this championship
    dp_stmt = (
        select(DraftPick)
        .join(Participant, DraftPick.participant_id == Participant.id)
        .where(Participant.championship_id == championship_id)
        .options(selectinload(DraftPick.participant).selectinload(Participant.team))
    )
    if participant_id:
        dp_stmt = dp_stmt.where(DraftPick.participant_id == participant_id)

    dp_result = await session.execute(dp_stmt)
    draft_picks = dp_result.scalars().all()

    # Build lookup: player_id -> (draft_pick, participant)
    dp_map = {}
    for dp in draft_picks:
        dp_map[dp.player_id] = (dp, dp.participant)

    responses = []
    for p in players:
        dp_info = dp_map.get(p.id)
        if participant_id and not dp_info:
            continue  # Filter: only show this participant's players
        dp = dp_info[0] if dp_info else None
        part = dp_info[1] if dp_info else None
        responses.append(_player_to_response(p, dp, part))

    return responses


@router.post("/{championship_id}/player/buy/{player_id}", response_model=PlayerResponse)
async def buy_player(
    championship_id: int,
    player_id: str,
    data: BuyPlayerModel,
    user: User = Depends(current_active_user),
    session: AsyncSession = Depends(get_async_session),
):
    # Verify championship ownership
    champ_stmt = select(Championship).where(
        Championship.id == championship_id, Championship.owner_id == str(user.id)
    )
    champ = (await session.execute(champ_stmt)).scalar_one_or_none()
    if not champ:
        raise HTTPException(status_code=404, detail="Championship not found.")

    # Get the player
    player_stmt = select(Player).options(
        selectinload(Player.position),
        selectinload(Player.team_origin_rel),
        selectinload(Player.nation_rel),
    ).where(Player.id == player_id)
    player = (await session.execute(player_stmt)).scalar_one_or_none()
    if not player:
        raise HTTPException(status_code=404, detail="Player not found.")

    # Check if player is already drafted in this championship
    existing_dp_stmt = (
        select(DraftPick)
        .join(Participant, DraftPick.participant_id == Participant.id)
        .where(
            Participant.championship_id == championship_id,
            DraftPick.player_id == player_id,
        )
        .options(selectinload(DraftPick.participant).selectinload(Participant.team))
    )
    existing_dp = (await session.execute(existing_dp_stmt)).scalar_one_or_none()

    # If player already drafted, refund and remove
    if existing_dp:
        old_participant = existing_dp.participant
        old_participant.budget += existing_dp.value
        await session.delete(existing_dp)
        await session.flush()

    # If no new participant specified, just undraft
    if not data.participant_id:
        await session.commit()
        return _player_to_response(player)

    if not data.value:
        raise HTTPException(status_code=400, detail="Player value not valid.")

    # Get new participant
    part_stmt = select(Participant).options(selectinload(Participant.team)).where(
        Participant.id == data.participant_id,
        Participant.championship_id == championship_id,
    )
    new_participant = (await session.execute(part_stmt)).scalar_one_or_none()
    if not new_participant:
        raise HTTPException(status_code=404, detail="Participant not found.")

    if new_participant.budget < data.value:
        raise HTTPException(status_code=400, detail="Participant does not have enough budget.")

    # Create draft pick and deduct budget
    new_participant.budget -= data.value
    new_dp = DraftPick(
        participant_id=new_participant.id,
        player_id=player_id,
        value=data.value,
    )
    session.add(new_dp)
    await session.commit()

    return _player_to_response(player, new_dp, new_participant)


@router.get("/{championship_id}/player/sheet")
async def get_sheet(
    championship_id: int,
    draft_only: bool = Query(True),
    user: User = Depends(current_active_user),
    session: AsyncSession = Depends(get_async_session),
):
    champ_stmt = select(Championship).where(
        Championship.id == championship_id, Championship.owner_id == str(user.id)
    )
    champ = (await session.execute(champ_stmt)).scalar_one_or_none()
    if not champ:
        raise HTTPException(status_code=404, detail="Championship not found.")

    positions_result = await session.execute(select(Position))
    positions = positions_result.scalars().all()

    p_count_stmt = select(func.count(Participant.id)).where(Participant.championship_id == championship_id)
    n_participants = (await session.execute(p_count_stmt)).scalar() or 0

    wb = openpyxl.Workbook()

    for i, pos in enumerate(positions):
        wb.create_sheet(index=i, title=pos.name)
        wb.active = i

        stmt = select(Player).options(
            selectinload(Player.position),
            selectinload(Player.team_origin_rel),
            selectinload(Player.nation_rel),
        ).where(Player.position_id == pos.id).order_by(Player.overall.desc(), Player.pace.desc())

        if draft_only:
            draft_names = _get_draft_player_ids(pos.name)
            if draft_names:
                stmt = stmt.where(Player.name.in_(draft_names))
            else:
                continue

        if n_participants and pos.name in _limit_by_position:
            stmt = stmt.limit(int(_limit_by_position[pos.name] * n_participants))

        result = await session.execute(stmt)
        players = result.scalars().all()

        # Get draft picks for this championship
        dp_stmt = (
            select(DraftPick)
            .join(Participant).where(Participant.championship_id == championship_id)
            .options(selectinload(DraftPick.participant))
        )
        dp_result = await session.execute(dp_stmt)
        dp_map = {dp.player_id: dp for dp in dp_result.scalars().all()}

        player_responses = []
        for p in players:
            dp = dp_map.get(p.id)
            part = dp.participant if dp else None
            player_responses.append(_player_to_response(p, dp, part))

        write_players_sheet(wb.active, player_responses)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return Response(
        content=stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sheet.xlsx"},
    )
