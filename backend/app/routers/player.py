# routes/users.py
import typing as t
import openpyxl
from io import BytesIO

from fastapi import APIRouter, HTTPException, Query, Response

from app.db.participant import get_participants, get_participants_num, update_participant_budget
from app.db.players import buy_player, get_player_by_id, get_player_transfers, get_players
from app.db.team import get_participant_id_by_team_name
from app.db.position import list_positions
from app.models.player import buy_player_model, player
from app.utils.export import write_players_sheet, write_transfers_sheet

_limit_by_position = {
    "Goalkeepers": 1,
    "Center Backs": 3,
    "Full Backs": 2.5,
    "Defensive Midfielders": 3,
    "Ofensive Midfielders": 1.5,
    "Wingers": 2.5,
    "Attackers": 2,
}

router = APIRouter()


@router.get("/list")
def list(
    position: str = Query(None, description="Filter by position"),
    team_participant: str = Query(None, description="Filter by team participant"),
) -> t.List[player]:
    where_clauses = []
    limit_clause = ""
    position_where_clause = ""

    if position:
        position_where_clause = f" WHERE id = {position}"

    positions_filtered = list_positions(where_clause=position_where_clause)
    
    if not len(positions_filtered):
        raise HTTPException(status_code=404, detail="Position not found.")
    
    position_obj = positions_filtered[0]

    if position:
        where_clauses.append(f"position_id = {position}")

        participants_number = get_participants_num()

        if participants_number:
            limit_clause = (
                f" LIMIT {int(participants_number * _limit_by_position[position_obj['name']])}"
            )

    if team_participant:
        where_clauses.append(f"team_participant = '{team_participant}'")

    return get_players(where_clauses, limit_clause)


@router.get("/sheet")
def get_sheet() -> t.Any:
    positions = list_positions()

    wb = openpyxl.Workbook()

    n_participants = get_participants_num()

    for i, position in enumerate(positions):
        wb.create_sheet(index=i, title=position["name"])
        wb.active = i

        where_clause = [f"position_id = {position['id']}"]
        limit_clause = f" LIMIT {int(_limit_by_position[position['name']] * n_participants)}"

        players = get_players(
            where_clauses=where_clause,
            limit_clause=limit_clause
        )

        write_players_sheet(wb.active, players)
    
    # Serialize the workbook to BytesIO
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    # Return the file-like object in response
    return Response(content=stream.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=sheet.xlsx"})


@router.get("/sheet-by-participant")
def get_participant_sheet() -> t.Any:
    participants = get_participants()

    wb = openpyxl.Workbook()

    for i, participant in enumerate(participants):
        wb.create_sheet(index=i, title=participant.name)
        wb.active = i

        where_clause = [f"team_participant = '{participant.team_name}'"]

        players = get_players(where_clauses=where_clause)

        write_players_sheet(wb.active, players)
    
    # Serialize the workbook to BytesIO
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    # Return the file-like object in response
    return Response(content=stream.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=sheet.xlsx"})


@router.post("/buy/{player_id}")
def buy(
    player_id: str,
    data: buy_player_model,
) -> player:
    player_instance = get_player_by_id(player_id)
    if not player_instance:
        raise HTTPException(status_code=404, detail="Player not found.")

    if player_instance.team_participant:
        # Update original participant budget and reset player's purchase info
        update_participant_budget(player_instance.participant_id, player_instance.value)
        buy_player(player_id)

    if not data.team_participant:
        return get_player_by_id(player_id)

    if not data.value:
        raise HTTPException(status_code=400, detail="Player value not valid")

    participant_id = get_participant_id_by_team_name(data.team_participant)
    update_participant_budget(participant_id, -data.value)
    buy_player(player_id, data.team_participant, data.value)

    return get_player_by_id(player_id)


@router.get("/transfers-sheet")
def get_transfers_sheet() -> t.Any:
    transfers = get_player_transfers()

    wb = openpyxl.Workbook()

    current_team_players = []
    current_team = None
    sheet_index = -1

    for player in transfers:
        if player["team_from"] != current_team:
            if current_team:
                write_transfers_sheet(wb.active, current_team_players)
                current_team_players = []

            current_team = player["team_from"]
            sheet_index += 1
            
            wb.create_sheet(index=sheet_index, title=current_team)
            wb.active = sheet_index

        current_team_players.append(player)

    # Serialize the workbook to BytesIO
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    # Return the file-like object in response
    return Response(content=stream.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=sheet.xlsx"})
