# db/crud.py

import os
import typing as t

import openpyxl

from app.models.player import player
from app.queries.players import (buy_player_query, change_players_team_query,
                                 player_exists_by_id_query,
                                 player_transfers_query, players_query,
                                 reset_players_team_by_participant_id_query)
from database.db import Database

database = Database()  # Initialize the custom database instance


def get_players(where_clauses: t.List[str], limit_clause: str = "") -> t.List[player]:
    where_clause_str = ""
    if where_clauses:
        where_clause_str = " WHERE " + " AND ".join(where_clauses)
    order_by_clause_str = " ORDER BY player.overall DESC, player.pace DESC"
    query = players_query + where_clause_str + order_by_clause_str + limit_clause
    results = database.execute_select_query(query)
    players = [player(**r) for r in results]

    return players


def get_player_by_id(player_id: str) -> player:
    where_clause = f" WHERE player.id = '{player_id}'"
    limit_clause = " LIMIT 1"
    query = players_query + where_clause + limit_clause
    results = database.execute_select_query(query)
    if not results:
        return None

    player_instance = player(**results[0])

    return player_instance


def check_player_exists(player_id: str) -> bool:
    args = {"player_id": player_id}
    result = database.execute_select_query(player_exists_by_id_query, args)

    return result[0]["player_count"] > 0


def buy_player(
    player_id: str,
    team_participant: t.Optional[str] = None,
    value: t.Optional[int] = None,
) -> None:
    args = (
        team_participant,
        value,
        player_id,
    )
    database.execute_query(buy_player_query, args)


def change_players_team(new_team_name: str, old_team_name: str) -> None:
    args = (
        new_team_name,
        old_team_name,
    )
    database.execute_query(change_players_team_query, args)


def reset_players_team_by_participant_id(participant_id: int) -> None:
    args = (participant_id,)
    database.execute_query(reset_players_team_by_participant_id_query, args)


def get_player_transfers() -> t.Any:
    results = database.execute_select_query(player_transfers_query)

    return results


def get_draft_player_ids(position_name: str = None) -> t.List[str]:
    """
    Read player names from sheet.xlsx file and convert them to player IDs.
    If position_name is provided, only read from that specific worksheet tab.
    Returns a list of player IDs that should be included in the draft.
    """
    try:
        # In Docker, the working directory is /app, so look for sheet.xlsx there
        # In local development, look in the backend directory
        possible_paths = [
            "/app/sheet.xlsx",  # Docker path
            os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                "sheet.xlsx",
            ),  # Local backend dir
            os.path.join(
                os.path.dirname(
                    os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
                ),
                "sheet.xlsx",
            ),  # Project root
        ]

        sheet_path = None
        for path in possible_paths:
            if os.path.exists(path):
                sheet_path = path
                break

        if not sheet_path:
            # If sheet.xlsx doesn't exist in any expected location, return empty list
            print("sheet.xlsx not found in any expected location")
            return []

        # Load the workbook
        workbook = openpyxl.load_workbook(sheet_path)

        player_names = []

        if position_name:
            # Try to find the worksheet with the position name
            worksheet = None
            for sheet_name in workbook.sheetnames:
                if sheet_name.lower() == position_name.lower():
                    worksheet = workbook[sheet_name]
                    break

            if not worksheet:
                print(f"Position sheet '{position_name}' not found in workbook")
                workbook.close()
                return []

            # Read player names from column A (starting from row 2 to skip header)
            for row in worksheet.iter_rows(min_row=2, max_col=1, values_only=True):
                if (
                    row[0] and str(row[0]).strip()
                ):  # If cell is not empty and not just whitespace
                    player_names.append(str(row[0]).strip())
        else:
            # If no position specified, read from all sheets
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                for row in worksheet.iter_rows(min_row=2, max_col=1, values_only=True):
                    if row[0] and str(row[0]).strip():
                        player_names.append(str(row[0]).strip())

        workbook.close()

        # Now convert player names to player IDs by querying the database
        if not player_names:
            return []

        # Escape single quotes in player names to prevent SQL injection
        escaped_names = [name.replace("'", "''") for name in player_names]
        names_str = "', '".join(escaped_names)
        query = f"SELECT id FROM `fifa-db`.player WHERE name IN ('{names_str}')"

        results = database.execute_select_query(query)
        player_ids = [str(result["id"]) for result in results]

        print(
            f"Found {len(player_names)} players in sheet{f' ({position_name})' if position_name else ''}, matched {len(player_ids)} in database"
        )
        return player_ids

    except Exception as e:
        # If any error occurs, log it and return empty list
        print(f"Error reading sheet.xlsx: {e}")
        return []
