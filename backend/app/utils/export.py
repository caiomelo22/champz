from openpyxl.styles import PatternFill


def colour_cell(cell, stat):
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    orange_fill = PatternFill(
        start_color="FF8F00", end_color="FF8F00", fill_type="solid"
    )

    yellow_fill = PatternFill(
        start_color="F6EB0A", end_color="F6EB0A", fill_type="solid"
    )

    light_green_fill = PatternFill(
        start_color="2DD63A", end_color="2DD63A", fill_type="solid"
    )

    green_fill = PatternFill(
        start_color="3E9C45", end_color="3E9C45", fill_type="solid"
    )

    cell.value = stat

    if stat >= 90:
        cell.fill = green_fill
    elif stat >= 80:
        cell.fill = light_green_fill
    elif stat >= 70:
        cell.fill = yellow_fill
    elif stat >= 60:
        cell.fill = orange_fill
    else:
        cell.fill = red_fill

    return cell


def write_players_sheet(sheet, players):
    sheet.oddHeader.center.size = 14
    sheet.oddHeader.center.font = "Calibri,Bold"

    sheet.cell(row=1, column=1).value = "PLAYER"
    sheet.cell(row=1, column=2).value = "TEAM"
    sheet.cell(row=1, column=3).value = "NATION"
    sheet.cell(row=1, column=4).value = "POSITION"
    sheet.cell(row=1, column=5).value = "OVERALL"
    sheet.cell(row=1, column=6).value = "PACE"
    sheet.cell(row=1, column=7).value = "SHOOTING"
    sheet.cell(row=1, column=8).value = "PASSING"
    sheet.cell(row=1, column=9).value = "DRIBBLING"
    sheet.cell(row=1, column=10).value = "DEFENDING"
    sheet.cell(row=1, column=11).value = "PHYSICAL"
    sheet.cell(row=1, column=12).value = "VALUE"

    for i, player in enumerate(players):
        sheet.cell(row=i + 2, column=1).value = player.name
        sheet.cell(row=i + 2, column=2).value = player.team_origin_name
        sheet.cell(row=i + 2, column=3).value = player.nation_name
        sheet.cell(row=i + 2, column=4).value = player.specific_position
        sheet.cell(row=i + 2, column=5).value = player.overall

        colour_cell(sheet.cell(row=i + 2, column=6), player.pace)
        colour_cell(sheet.cell(row=i + 2, column=7), player.shooting)
        colour_cell(sheet.cell(row=i + 2, column=8), player.passing)
        colour_cell(sheet.cell(row=i + 2, column=9), player.dribbling)
        colour_cell(sheet.cell(row=i + 2, column=10), player.defending)
        colour_cell(sheet.cell(row=i + 2, column=11), player.physical)

        if player.value:
            sheet.cell(row=i + 2, column=12).value = player.value

    auto_size_columns(sheet)

    return sheet


def write_transfers_sheet(sheet, players):
    sheet.oddHeader.center.size = 14
    sheet.oddHeader.center.font = "Calibri,Bold"

    sheet.cell(row=1, column=1).value = "PLAYER"
    sheet.cell(row=1, column=2).value = "OVERALL"
    sheet.cell(row=1, column=3).value = "TO"

    for i, player in enumerate(players):
        sheet.cell(row=i + 2, column=1).value = player["name"]
        sheet.cell(row=i + 2, column=2).value = player["overall"]
        sheet.cell(row=i + 2, column=3).value = player["team_to"]

    auto_size_columns(sheet)

    return sheet


def auto_size_columns(sheet):
    # Auto-size columns
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (
            max_length + 2
        ) * 1.2  # Add some padding and adjust for font width
        sheet.column_dimensions[column].width = adjusted_width
